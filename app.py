import os
import datetime
import segno
from flask import Flask, request, jsonify, render_template, send_file, send_from_directory
from werkzeug.utils import secure_filename
from openpyxl import Workbook, load_workbook
from ultralytics import YOLO

app = Flask(__name__)


UPLOAD_DIR = 'uploads'
QR_DIR = 'qrcodes'
EXCEL_FILE = 'medical_waste_records.xlsx'
MODEL_PATH = 'best.pt'

# 类别名称，注意索引顺序要与训练时一致
CLASS_NAMES = ['化学性废物', '锐器', '感染性废物', '病理性废物']


if not os.path.exists(UPLOAD_DIR):
    os.makedirs(UPLOAD_DIR)
if not os.path.exists(QR_DIR):
    os.makedirs(QR_DIR)

app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB，够用了


_model = None

def load_model():
    global _model
    if _model is None:
        if not os.path.exists(MODEL_PATH):
            raise FileNotFoundError(f"模型文件 {MODEL_PATH} 没找到")
        _model = YOLO(MODEL_PATH)
        print("模型加载完成")
    return _model


def init_excel():
    """如果Excel文件不存在创建，加上表头"""
    if not os.path.exists(EXCEL_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "医疗垃圾记录"
        headers = ['记录编号', '垃圾类别', '具体类型', '置信度', '重量(g)', '产生科室', '操作员', '处理时间', '二维码文件']
        ws.append(headers)
        wb.save(EXCEL_FILE)
        print("数据库文件已创建")
    else:
        # 简单检查表头是否完整
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        if ws.max_row == 0 or ws.cell(row=1, column=1).value != '记录编号':
            print("Excel格式错误，重新生成表头")
            ws.delete_rows(1, ws.max_row)
            headers = ['记录编号', '垃圾类别', '具体类型', '置信度', '重量(g)', '产生科室', '操作员', '处理时间', '二维码文件']
            ws.append(headers)
            wb.save(EXCEL_FILE)


def save_one_record(record):
    try:
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
        ws.append(record)
        wb.save(EXCEL_FILE)
        return True
    except Exception as e:
        print("保存Excel失败：", e)
        return False

def predict_image(image_path):
    model = load_model()
    results = model(image_path)
    best = None
    for r in results:
        if r.boxes is None:
            continue
        for box in r.boxes:
            cls_id = int(box.cls[0])
            conf = float(box.conf[0])
            name = CLASS_NAMES[cls_id] if cls_id < len(CLASS_NAMES) else f"未知类别({cls_id})"
            if best is None or conf > best['confidence']:
                best = {
                    'category': name,
                    'confidence': conf,
                    'bbox': box.xyxy[0].tolist()
                }
    return best


@app.route('/')
def index():
    return render_template('index.html')

@app.route('/export', methods=['GET'])
def export():
    if not os.path.exists(EXCEL_FILE):
        return jsonify({'error': '没有数据'}), 404
    export_name = f"export_{datetime.datetime.now().strftime('%Y%m%d%H%M%S')}.xlsx"
    wb = load_workbook(EXCEL_FILE)
    wb.save(export_name)
    return send_file(export_name, as_attachment=True, download_name=export_name)

@app.route('/qrcodes/<filename>')
def qrcode_file(filename):
    return send_from_directory(QR_DIR, filename)

@app.route('/upload', methods=['POST'])
def upload():
    if 'image' not in request.files:
        return jsonify({'error': '没有图片文件'}), 400
    f = request.files['image']
    if f.filename == '':
        return jsonify({'error': '文件名为空'}), 400

    weight = request.form.get('weight', '0.0')
    dept = request.form.get('department', '未指定')
    operator = request.form.get('operator', '系统')

    filename = secure_filename(f.filename)
    filepath = os.path.join(UPLOAD_DIR, filename)
    f.save(filepath)

    try:
        result = predict_image(filepath)
        if result is None:
            return jsonify({'error': '没有识别到医疗垃圾'}), 400
    except Exception as e:
        return jsonify({'error': f'识别出错: {str(e)}'}), 500

    category = result['category']
    conf = result['confidence']

    record_id = datetime.datetime.now().strftime("%Y%m%d%H%M%S")
    now_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S')


    qr_content = f"""医疗废物记录
编号：{record_id}
类别：{category}
置信度：{conf:.2%}
重量：{weight}g
科室：{dept}
操作员：{operator}
时间：{now_time}"""

    qr_filename = f"{record_id}.png"
    qr_path = os.path.join(QR_DIR, qr_filename)
    try:
        qr = segno.make(qr_content)
        qr.save(qr_path, scale=10)
    except Exception as e:
        return jsonify({'error': f'二维码生成失败: {str(e)}'}), 500


    record_data = [
        record_id,
        category,
        '',
        conf,
        weight,
        dept,
        operator,
        now_time,
        qr_filename
    ]

    if not record_data[1]:
        return jsonify({'error': '类别为空，保存失败'}), 500
    if not save_one_record(record_data):
        return jsonify({'error': '数据保存失败'}), 500

    return jsonify({
        'record_id': record_id,
        'category': category,
        'confidence': conf,
        'qr_file': qr_filename,
        'time': now_time,
        'weight': weight,
        'department': dept,
        'operator': operator
    })


if __name__ == '__main__':
    init_excel()
    try:
        load_model()
    except Exception as e:
        print("模型加载失败，请检查路径", e)
    app.run(debug=True, host='0.0.0.0', port=5000)